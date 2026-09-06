# -*- coding: utf-8 -*-
"""
dispatch_driver_baecha_to_excel.py

sqlite3의 dispatch_daily 테이블에서 데이터를 읽어, 이미지와 같은 형식(사번/성명/휴무일 +
일자별 근무표 + 오전/오후/정상 집계)의 엑셀 파일을 생성한다.

사용법:
    python dispatch_driver_baecha_to_excel.py --db dispatch.db --start 2026-08-20 --out 배차표.xlsx
    python dispatch_driver_baecha_to_excel.py --db dispatch.db --start 2026-09-06 --weeks 4

테이블 스키마 
    CREATE TABLE IF NOT EXISTS dispatch_daily (
        month         TEXT NOT NULL,
        name          TEXT NOT NULL,
        day           INTEGER,
        employee_no   TEXT NOT NULL,
        weekday       TEXT,
        value         TEXT,
        bg_color      TEXT,
        created_at    TEXT DEFAULT (datetime('now','localtime')),
        PRIMARY KEY(month, name, day)
    );
"""

import argparse
from calendar import weekday
import datetime
import sqlite3
from collections import defaultdict, Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -------------------------------------------------------------------------
# 기본 설정 (필요시 조정)
# -------------------------------------------------------------------------
FONT_NAME = "맑은 고딕"          # 한글 표기가 있으므로 맑은 고딕 사용 (Arial은 한글 미지원)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")      # 상단 고정 컬럼 헤더
DAY_HEADER_FILL = PatternFill("solid", fgColor="E2EFDA")  # 일자 컬럼 헤더
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")

WEEKDAY_KR = ["월", "화", "수", "목", "금", "토", "일"]  # datetime.weekday(): 0=월


def to_argb(hex_color):
    """'#ffff00' -> 'FFFFFF00' (openpyxl은 ARGB 8자리를 요구)."""
    if not hex_color:
        return None
    h = hex_color.lstrip("#")
    if len(h) == 6:
        return "FF" + h.upper()
    if len(h) == 8:
        return h.upper()
    return None


def build_date_list(start_date, weeks):
    """start_date부터 weeks*7일 간의 date 객체 리스트."""
    return [start_date + datetime.timedelta(days=i) for i in range(weeks * 7)]


def fetch_records(conn, dates):
    """
    date 리스트에 해당하는 (month, day) 조합으로 dispatch_daily를 조회한다.
    월 경계를 넘어가는 경우(예: 4주가 두 달에 걸침)도 함께 처리한다.

    반환: {(name, date): {"name":..., "value":..., "bg_color":...}}
    그리고 사번->성명 매핑, 사번 정렬 리스트
    """
    # 월별로 (month, day) 그룹핑
    by_month = defaultdict(list)
    for d in dates:
        qdate = d.strftime("%Y-%m")
        by_month[qdate].append(d.day)

    records = {}  # (name, date) -> row dict
    names = {}    # employee_no -> name

    cur = conn.cursor()
    for qdate, days in by_month.items():
        placeholders = ",".join("?" for _ in days)
        sql = f"""
            SELECT name, day, weekday, value, bg_color
            FROM dispatch_daily
            WHERE month = ? AND day IN ({placeholders})
        """
        cur.execute(sql, [qdate, *days])
        for name, day, weekday, value, bg_color in cur.fetchall():
            d = datetime.date(int(qdate[:4]), int(qdate[5:7]), day)
            records[(name, d)] = {
                "value": value or "",
                "bg_color": bg_color,
            }
            names[name] = name

    name_nos = sorted(names.keys())
    return records, names, name_nos


def guess_day_off(records, name, dates):
    """
    [참고용 추정치] dispatch_daily 테이블에는 '고정 휴무 요일'을 담는 컬럼이 없다.
    같은 요일에 값이 비어있는(공백/None) 경우가 가장 많은 요일을 '휴무일' 후보로 추정한다.
    실제 고정휴무 데이터가 별도 테이블/컬럼에 있다면 이 함수 대신 그 값을 사용할 것.
    """
    counter = Counter()
    total = Counter()
    for d in dates:
        row = records.get((name, d))
        wd = WEEKDAY_KR[d.weekday()]
        total[wd] += 1
        if row and not row["value"].strip():
            counter[wd] += 1
    if not counter:
        return ""
    # 가장 빈 값이 많이 나온 요일 (동률이면 요일 순서상 먼저 나오는 것)
    best_wd = max(counter, key=lambda wd: (counter[wd], -WEEKDAY_KR.index(wd)))
    return best_wd if counter[best_wd] >= 1 else ""


def build_workbook(records, names, name_nos, dates, include_day_off_guess=False):
    wb = Workbook()
    ws = wb.active
    ws.title = "배차표"

    # ---------------- 헤더 ----------------
    fixed_headers = ["성명", "휴무"]
    tail_headers = ["오전", "오후", "정상"]
    n_fixed = len(fixed_headers)
    n_days = len(dates)

    for col, title in enumerate(fixed_headers, start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = Font(name=FONT_NAME, bold=True)
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    for i, d in enumerate(dates):
        col = n_fixed + 1 + i
        wd = WEEKDAY_KR[d.weekday()]
        c = ws.cell(row=1, column=col, value=f"{d.day:02d} {wd}")
        c.fill = DAY_HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
        if wd == "일":
            c.font = Font(name=FONT_NAME, bold=True, color="FF0000")   # 일요일: 빨강+굵게
        elif wd == "토":
            c.font = Font(name=FONT_NAME, bold=True, color="0070C0")   # 토요일: 파랑+굵게
        else:
            c.font = Font(name=FONT_NAME, bold=True)

    tail_start_col = n_fixed + n_days + 1
    for i, title in enumerate(tail_headers):
        c = ws.cell(row=1, column=tail_start_col + i, value=title)
        c.font = Font(name=FONT_NAME, bold=True)
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    # ---------------- 데이터 ----------------
    day_col_first = n_fixed + 1
    day_col_last = n_fixed + n_days

    for r, name in enumerate(name_nos, start=2):
        ws.cell(row=r, column=1, value=names.get(name, "")).border = BORDER

        #day_off = guess_day_off(records, name, dates) if include_day_off_guess else ""

        day_off_from_bg = ""

        for i, d in enumerate(dates):
            col = day_col_first + i
            row = records.get((name, d), {"value": "", "bg_color": None})
            cell = ws.cell(row=r, column=col, value=row["value"] or None)
            cell.alignment = CENTER
            cell.border = BORDER
            cell.font = Font(name=FONT_NAME)
            argb = to_argb(row["bg_color"])
            if argb:
                cell.fill = PatternFill("solid", fgColor=argb)
                if row["bg_color"] == "#ff94dc" and day_off_from_bg == "":
                    words = ws.cell(row=1,column=col).value.split()
                    if len(words) > 1:
                        day_off_from_bg = words[1]
                #'print(f"추정 휴무일: {emp_no} {names.get(emp_no)} {day_off_from_bg} ({row['bg_color']})")

        ws.cell(row=r, column=2, value=day_off_from_bg).border = BORDER

        # 오전 / 오후 / 정상 은 COUNTIF 수식으로 작성 (openpyxl 하드코딩 금지 원칙)
        first_letter = get_column_letter(day_col_first)
        last_letter = get_column_letter(day_col_last)
        rng = f"{first_letter}{r}:{last_letter}{r}"

        c_am = ws.cell(row=r, column=tail_start_col, value=f'=COUNTIF({rng},"A")')
        c_pm = ws.cell(row=r, column=tail_start_col + 1, value=f'=COUNTIF({rng},"P")')
        c_total = ws.cell(
            row=r,
            column=tail_start_col + 2,
            value=f"={get_column_letter(tail_start_col)}{r}+{get_column_letter(tail_start_col + 1)}{r}",
        )
        for c in (c_am, c_pm, c_total):
            c.alignment = CENTER
            c.border = BORDER
            c.font = Font(name=FONT_NAME, bold=True)

    # ---------------- 열 너비 / 틀고정 ----------------
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 8
    for i in range(n_days):
        ws.column_dimensions[get_column_letter(day_col_first + i)].width = 6
    for i in range(3):
        ws.column_dimensions[get_column_letter(tail_start_col + i)].width = 7

    ws.freeze_panes = ws.cell(row=2, column=day_col_first)
    ws.row_dimensions[1].height = 20

    return wb


def main():
    ap = argparse.ArgumentParser(description="dispatch_daily -> 4주 배차표 엑셀 생성")
    ap.add_argument("--db", required=True, help="sqlite db 파일 경로")
    ap.add_argument("--start", required=True, help="시작일자 YYYY-MM-DD")
    ap.add_argument("--weeks", type=int, default=4, help="추출할 주 수 (기본 4주)")
    ap.add_argument("--out", default=None, help="출력 엑셀 경로 (기본: 배차표_시작일_N주.xlsx)")
    ap.add_argument(
        "--guess-day-off",
        action="store_true",
        help="휴무일 컬럼을 값이 비어있는 요일로 추정해서 채움 (DB에 실제 컬럼이 없을 때의 참고용 추정치)",
    )
    args = ap.parse_args()

    start_date = datetime.datetime.strptime(args.start, "%Y-%m-%d").date()
    dates = build_date_list(start_date, args.weeks)
    end_date = dates[-1]

    out_path = args.out or f"배차표_{start_date.isoformat()}_{args.weeks}주.xlsx"

    conn = sqlite3.connect(args.db)
    try:
        records, names, name_nos = fetch_records(conn, dates)
    finally:
        conn.close()

    if not name_nos:
        print(f"경고: {start_date} ~ {end_date} 기간에 해당하는 데이터가 없습니다.")

    wb = build_workbook(records, names, name_nos, dates, include_day_off_guess=args.guess_day_off)
    wb.save(out_path)
    print(f"완료: {out_path} ({len(name_nos)}명, {start_date} ~ {end_date})")


if __name__ == "__main__":
    main()
