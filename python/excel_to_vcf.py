import os
import quopri
import sys
from pathlib import Path

import openpyxl

# ============================================================
# 설정
# ============================================================

# 입력 Excel 파일명
INPUT_XLSX = "..\\전화번호\\연락처조정작업_2.xlsm"

# 출력 VCF 파일명
OUTPUT_VCF = "..\\전화번호\\new_contacts.vcf"

# 이 스크립트 파일이 있는 폴더로 작업 디렉토리를 강제 고정 (디버그 실행시 작업디렉토리를 환경파일 폴더로 한다. 이를 소스폴더로 변경하려면...)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
os.chdir(SCRIPT_DIR)   # -> 현재소스의 경로를 cwd 고정


# ============================================================
# UTF-8 문자열 -> Quoted-Printable
#   -> =EA=B0=95=EC=A0=95...
# ============================================================

def qp(text):
    if text is None:
        return ""
    text = str(text)
    return quopri.encodestring(
        text.encode("utf-8"),
        quotetabs=True
    ).decode("ascii")


# ============================================================
# Excel -> VCF
# ============================================================

def make_vcf(param_order, output_vcf, input_xlsx ):
    wb = openpyxl.load_workbook(input_xlsx, data_only=True)
    ws = wb.worksheets[0] #wb.active

    # 첫 번째 행을 헤더로 사용
    headers = {}
    for cell in ws[1]:
        if cell.value is not None:
            headers[str(cell.value).strip()] = cell.column

    required = ["tel_section", "tel_number", "fn_new","fn_order","fn_old"]

    for name in required:
        if name not in headers:
            raise ValueError(
                f"Excel에 '{name}' 열이 없습니다. "
                f"현재 열: {list(headers.keys())}"
            )

    tel_section_col = headers["tel_section"]
    tel_number_col = headers["tel_number"]
    fn_new_col = headers["fn_new"]
    fn_order_col = headers["fn_order"]
    fn_old_col = headers["fn_old"]
    vcards = []

    for row in range(2, ws.max_row + 1):

        if param_order != "all" and ws.cell(row, fn_order_col).value != param_order:
            continue

        tel_section = ws.cell(row, tel_section_col).value
        tel_number = ws.cell(row, tel_number_col).value
        fn_new = ws.cell(row, fn_new_col).value
        fn_old = ws.cell(row, fn_old_col).value

        # 빈 행은 건너뜀
        if tel_number is None:
            continue
        if fn_new is None:
            fn_new = "ㅎㅎㅎ삭제대상_" + str(fn_old)

        tel_section = str(tel_section).strip() if tel_section is not None else "cell"
        fn_new = str(fn_new).strip()

        # Excel에서 전화번호가 숫자로 읽힌 경우를 대비
        if isinstance(tel_number, float) and tel_number.is_integer():
            tel_number = str(int(tel_number))
        else:
            tel_number = str(tel_number).strip()

        # 전화번호 앞의 0이 Excel에서 사라진 경우에는
        # 여기서 자동 복구할 수 없으므로 Excel 열을 '텍스트'로
        # 저장하는 것을 권장합니다.

        # tel_section -> VCF TEL type
        # 예: cell -> CELL
        tel_type = tel_section.upper()

        encoded_name = qp(fn_new)

        vcard = (
            "BEGIN:VCARD\r\n"
            "VERSION:2.1\r\n"
            f"N;CHARSET=UTF-8;ENCODING=QUOTED-PRINTABLE:;{encoded_name};;;\r\n"
            f"FN;CHARSET=UTF-8;ENCODING=QUOTED-PRINTABLE:{encoded_name}\r\n"
            f"TEL;{tel_type}:{tel_number}\r\n"
            "END:VCARD\r\n"
        )

        vcards.append(vcard)

    # UTF-8이지만 VCF 자체의 QP 데이터는 ASCII이므로
    # UTF-8로 저장해도 갤럭시에서 읽을 수 있습니다.
    Path(output_vcf).write_text(
        "".join(vcards),
        encoding="utf-8",
        newline=""
    )

    print(f"완료: {output_vcf}")
    print(f"생성된 연락처 수: {len(vcards)}")


if __name__ == "__main__":
    param_order = sys.argv[1] if len(sys.argv) > 1 else "all"
    output_vcf = sys.argv[2] if len(sys.argv) > 2 else OUTPUT_VCF
    input_xlsx = sys.argv[3] if len(sys.argv) > 3 else INPUT_XLSX
    make_vcf(param_order, output_vcf, input_xlsx )
