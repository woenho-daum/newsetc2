"""
dispatch.db(SQLite)의 dispatch_settings 테이블에서
사번(driver_no)을 매칭시켜 route(노선), car_number(차량번호), apply_date(적용일자)를
엑셀 파일의 AH, AI, AJ 컬럼에 채워 넣는 스크립트.

사용법:
    py .\fill_excel_ from_dispatch_driver_baecha.py --excel_path .\배차표.xlsx --db_path .\baecha.db
    python fill_excel_ from_dispatch_driver_baecha.py <엑셀파일.xlsx> <dispatch.db> [--sheet 시트명] [--name-col A] [--start-row 2]
기본값:
    - 시트: 활성 시트(맨 앞 시트) 사용
    - 성명 컬럼: A열
    - 데이터 시작행: 2행 (1행은 헤더로 간주)
    - 출력 컬럼: AH(route), AI(car_number), AJ(apply_date)

주의:
    - dispatch_settings 테이블은 driver_no(사번) 별로 A/B 두 행이 있을 수 있으므로
      기본적으로 각 driver_no당 가장 최근(collected_at 최댓값) 레코드 1건만 사용한다.
      (여러 행을 모두 반영하고 싶다면 아래 --keep-all-rows 옵션을 쓰면 A/B를 이어붙여 표기한다)
"""

import argparse
import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


def load_dispatch_map(db_path: str, keep_all_rows: bool = False):
    """
    dispatch_settings 테이블을 읽어서
    { driver_no: {"route": ..., "car_number": ..., "apply_date": ...} } 형태의 dict를 만든다.

    같은 driver_no가 여러 건(A/B 등) 있을 경우:
      - keep_all_rows=False (기본): collected_at이 가장 최신인 행 1건만 사용
      - keep_all_rows=True: route/car_number/apply_date를 " / "로 이어붙여서 표기
    """
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute(
        """
        SELECT driver_name, driver_type, route, car_number, apply_date, collected_at
        FROM dispatch_settings
        ORDER BY driver_name, collected_at DESC
        """
    )
    rows = cur.fetchall()
    conn.close()

    result = {}
    for row in rows:
        dno = row["driver_name"]
        if dno is None:
            continue
        dno = str(dno).strip()
        if not dno:
            continue

        entry = {
            "route": row["route"] or "",
            "car_number": row["car_number"] or "",
            "apply_date": row["apply_date"] or "",
        }

        if dno not in result:
            result[dno] = entry
        elif keep_all_rows:
            existing = result[dno]
            for key in ("route", "car_number", "apply_date"):
                new_val = entry[key]
                if new_val and new_val not in existing[key].split(" / "):
                    existing[key] = f"{existing[key]} / {new_val}" if existing[key] else new_val
        # keep_all_rows=False 인 경우, ORDER BY collected_at DESC 이므로
        # 처음 만난(=가장 최신) 행만 유지하고 이후 행은 무시한다.

    return result


def fill_excel(
    excel_path: str,
    dispatch_map: dict,
    sheet_name: str | None,
    name_col: str,
    route_col: str,
    car_col: str,
    date_col: str,
    start_row: int,
):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    name_idx = column_index_from_string(name_col)
    route_idx = column_index_from_string(route_col)
    car_idx = column_index_from_string(car_col)
    date_idx = column_index_from_string(date_col)

    matched = 0
    unmatched = []

    for r in range(start_row, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=name_idx).value
        if cell_val is None or str(cell_val).strip() == "":
            continue

        driver_no = str(cell_val).strip()
        # 사번이 "00099" 처럼 앞자리 0이 있는 경우, 엑셀에서 숫자로 읽혀
        # "99" 로 들어올 수 있어 0-padding도 함께 시도한다.
        info = dispatch_map.get(driver_no)
        if info is None:
            # 5자리 기준으로 zfill 시도 (필요시 조정)
            padded = driver_no.zfill(5)
            info = dispatch_map.get(padded)

        if info is None:
            unmatched.append(driver_no)
            continue

        ws.cell(row=r, column=route_idx, value=info["route"])
        ws.cell(row=r, column=car_idx, value=info["car_number"])
        ws.cell(row=r, column=date_idx, value=info["apply_date"])
        matched += 1

    wb.save(excel_path)
    return matched, unmatched


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--excel_path", help="대상 엑셀 파일 경로 (.xlsx)")
    parser.add_argument("--db_path", help="dispatch.db(SQLite) 경로")
    parser.add_argument("--sheet", default=None, help="시트명 (생략 시 활성 시트)")
    parser.add_argument("--name-col", default="A", help="성명이 들어있는 컬럼 (기본 A)")
    parser.add_argument("--route-col", default="AH", help="노선 출력 컬럼 (기본 AI)")
    parser.add_argument("--car-col", default="AI", help="차량번호 출력 컬럼 (기본 AJ)")
    parser.add_argument("--date-col", default="AJ", help="적용일자 출력 컬럼 (기본 AK)")
    parser.add_argument("--start-row", type=int, default=2, help="데이터 시작 행 (기본 2, 1행은 헤더)")
    parser.add_argument(
        "--keep-all-rows",
        action="store_true",
        help="driver_no별로 여러 행(A/B 등)이 있을 때 모두 이어붙여 표기 (기본은 최신 1건만 사용)",
    )
    args = parser.parse_args()

    if not Path(args.excel_path).exists():
        print(f"[오류] 엑셀 파일을 찾을 수 없습니다: {args.excel_path}", file=sys.stderr)
        sys.exit(1)
    if not Path(args.db_path).exists():
        print(f"[오류] DB 파일을 찾을 수 없습니다: {args.db_path}", file=sys.stderr)
        sys.exit(1)

    print(f"[정보] DB에서 dispatch_settings 로딩 중: {args.db_path}")
    dispatch_map = load_dispatch_map(args.db_path, keep_all_rows=args.keep_all_rows)
    print(f"[정보] driver_no {len(dispatch_map)}건 로딩 완료")

    print(f"[정보] 엑셀 파일 처리 중: {args.excel_path}")
    matched, unmatched = fill_excel(
        excel_path=args.excel_path,
        dispatch_map=dispatch_map,
        sheet_name=args.sheet,
        name_col=args.name_col,
        route_col=args.route_col,
        car_col=args.car_col,
        date_col=args.date_col,
        start_row=args.start_row,
    )

    print(f"[완료] 매칭 성공: {matched}건")
    if unmatched:
        print(f"[경고] 매칭 실패(성명 없음): {len(unmatched)}건")
        print("  -> " + ", ".join(unmatched[:30]) + (" ..." if len(unmatched) > 30 else ""))
    print(f"[완료] 저장됨: {args.excel_path}")


if __name__ == "__main__":
    main()
