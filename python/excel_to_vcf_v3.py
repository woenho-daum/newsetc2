import os
import sys
from pathlib import Path

import openpyxl

# ============================================================
# 설정
# ============================================================

# 입력 Excel 파일명
INPUT_XLSX = "..\\전화번호\\연락처조정작업_2.xlsm"

# 출력 VCF 파일명
OUTPUT_VCF = "..\\전화번호\\new_contacts.vcf"

# 이 스크립트 파일이 있는 폴더로 작업 디렉토리를 강제 고정 (디버그 실행시 작업디렉토리를 환경파일 폴더로 한다. 이를 소스폴더로 변경하려면...)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
os.chdir(SCRIPT_DIR)   # -> 현재소스의 경로를 cwd 고정


# ============================================================
# vCard 3.0 / UTF-8
# ============================================================

# 전화번호 앞의 0이 Excel에서 사라진 경우에는
# 여기서 자동 복구할 수 없으므로 Excel 열을 '텍스트'로
# 저장하는 것을 권장합니다.

def get_excel_data(value):
    if isinstance(value, str):
        # value가 문자열일 때만 실행
        return str(value).strip()
    elif isinstance(value, int):
        # value가 정수일 때만 실행
        return str(value)
    else:
        return ""

# ============================================================
# Excel -> VCF
# ============================================================

def make_vcf(param_order, output_vcf, input_xlsx ):
    wb = openpyxl.load_workbook(input_xlsx, data_only=True)
    ws = wb.worksheets[0] #wb.active

    # 첫 번째 행을 헤더로 사용
    headers = {}
    for cell in ws[1]:
        if cell.value is not None:
            headers[str(cell.value).strip()] = cell.column

    required = ["tel_section", "tel_number", "fn_new","fn_order","fn_old"]

    for name in required:
        if name not in headers:
            raise ValueError(
                f"Excel에 '{name}' 열이 없습니다. "
                f"현재 열: {list(headers.keys())}"
            )

    tel_section_col = headers["tel_section"]
    tel_number_col = headers["tel_number"]
    fn_new_col = headers["fn_new"]
    fn_order_col = headers["fn_order"]
    fn_old_col = headers["fn_old"]
    vcards = []

    for row in range(2, ws.max_row + 1):

        # param_order이 "all"이 아니고, 현재 행의 fn_order 값이 param_order과 다르면 건너뜀
        if param_order != "all" and ws.cell(row, fn_order_col).value != param_order:
            continue

        tel_section = get_excel_data(ws.cell(row, tel_section_col).value)
        tel_number = get_excel_data(ws.cell(row, tel_number_col).value)
        fn_new = get_excel_data(ws.cell(row, fn_new_col).value)
        fn_old = get_excel_data(ws.cell(row, fn_old_col).value)
        fn_order = get_excel_data(ws.cell(row, fn_order_col).value)

        # 빈 행은 건너뜀
        #if tel_number in ("", None) or fn_old in ("", None) or fn_order == "삭제":
        if tel_number == "" or fn_old == "" or fn_order == "삭제":
            continue

        if fn_new == "":
            fn_new = "힣힣힣삭제대상_" + str(fn_old)

        #if tel_number=="01053793822":
        #    print(f"디버그1: tel_number={tel_number}, fn_new={fn_new}, fn_old={fn_old}")
        
        if tel_section == "":
            tel_section = "cell"

        #if tel_number=="01053793822":
        #    print(f"디버그2: tel_number={tel_number}, fn_new={fn_new}, fn_old={fn_old}")

        # tel_section -> VCF TEL type
        # 예: cell -> CELL
        tel_type = tel_section.upper()

        encoded_name = fn_new

        if encoded_name in (None, ""):
            print(f"경고: tel_number={tel_number}, fn_new={fn_new}, fn_old={fn_old}, encoded_name={encoded_name}")
            continue


        # Google 주소록용 그룹:
        # fn_order가 "고정/쉬프트"이면 fn_new의 공백 기준 2번째 단어를 CATEGORIES로 사용
        category = ""
        if fn_order == "고정/쉬프트":
            words = fn_new.split()
            if len(words) > 1:
                category = words[1]
        else:
            category = ""

        vcard = (
            "BEGIN:VCARD\r\n"
            "VERSION:3.0\r\n"
            f"N;CHARSET=UTF-8:;{encoded_name};;;\r\n"
            f"FN;CHARSET=UTF-8:{encoded_name}\r\n"
            f"TEL;TYPE={tel_type}:{tel_number}\r\n"
            + (f"CATEGORIES:{category}\r\n" if category else "")
            + "END:VCARD\r\n"
        )

        vcards.append(vcard)

    # vCard 3.0을 UTF-8 그대로 저장
    Path(output_vcf).write_text(
        "".join(vcards),
        encoding="utf-8",
        newline=""
    )

    print(f"완료: {output_vcf}")
    print(f"생성된 연락처 수: {len(vcards)}")


if __name__ == "__main__":
    param_order = sys.argv[1] if len(sys.argv) > 1 else "all"
    output_vcf = sys.argv[2] if len(sys.argv) > 2 else OUTPUT_VCF
    input_xlsx = sys.argv[3] if len(sys.argv) > 3 else INPUT_XLSX
    make_vcf(param_order, output_vcf, input_xlsx )
